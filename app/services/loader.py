from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import unicodedata

from openpyxl import load_workbook

from app.models import Titulo


class DataFileError(Exception):
    pass


def _norm(s: Any) -> str:
    """Normaliza cabeçalho: lower, remove acentos, remove pontuação/espaços."""
    if s is None:
        return ""
    txt = str(s).strip().lower()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    txt = "".join(ch for ch in txt if ch.isalnum())
    return txt


SYNONYMS = {
    "emissor": ["emissor", "instituicao", "banco"],
    "tipo": ["tipo", "produto", "titulo"],
    "vencimento": ["vencimento", "datavencimento"],
    "taxa": ["txportal", "taxadoportalas10h1", "taxadoportalas10h", "taxadoportal", "taxa", "txmaxima"],
}


def _pick_col(colmap: Dict[str, int], canonical: str) -> Optional[int]:
    for syn in SYNONYMS[canonical]:
        if syn in colmap:
            return colmap[syn]
    return None


def _find_header_and_map(ws, max_scan_rows: int = 50) -> Optional[Tuple[int, Dict[str, int]]]:
    """
    Procura uma linha de cabeçalho que contenha pelo menos: tipo + vencimento + taxa.
    Retorna (header_row, colmap_norm->idx) ou None.
    """
    best: Optional[Tuple[int, Dict[str, int], int]] = None

    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        headers = [_norm(c.value) for c in ws[r]]
        colmap = {h: i for i, h in enumerate(headers) if h}

        score = 0
        for key in ("tipo", "vencimento", "taxa", "emissor"):
            if _pick_col(colmap, key) is not None:
                score += 1

        if score >= 3:
            if best is None or score > best[2]:
                best = (r, colmap, score)

    if best is None:
        return None
    return best[0], best[1]


def _row_to_titulo(
    row: Tuple[Any, ...],
    colmap: Dict[str, int],
    sheet_name: str,
) -> Optional[Dict[str, Any]]:
    idx_tipo = _pick_col(colmap, "tipo")
    idx_venc = _pick_col(colmap, "vencimento")
    idx_taxa = _pick_col(colmap, "taxa")
    idx_emis = _pick_col(colmap, "emissor")

    if idx_tipo is None or idx_venc is None or idx_taxa is None:
        return None

    tipo = row[idx_tipo] if idx_tipo < len(row) else None
    venc = row[idx_venc] if idx_venc < len(row) else None
    taxa = row[idx_taxa] if idx_taxa < len(row) else None
    emissor = row[idx_emis] if (idx_emis is not None and idx_emis < len(row)) else None

    if emissor is None and "titul" in _norm(sheet_name):
        emissor = "Tesouro Nacional"

    if not tipo or not venc or taxa is None or not emissor:
        return None

    return {
        "tipo": tipo,
        "vencimento": venc,
        "taxa": taxa,
        "emissor": emissor,
    }


def load_titulos(path_str: str, sheet_name: Optional[str] = None) -> List[Titulo]:
    path = Path(path_str)
    if not path.exists():
        raise DataFileError(f"Arquivo não encontrado: {path}")

    ext = path.suffix.lower()
    if ext not in {".xlsm", ".xlsx"}:
        raise DataFileError(f"Extensão não suportada: {ext}. Use .xlsm/.xlsx")

    wb = load_workbook(filename=path, data_only=True, keep_vba=True)

    sheets = [wb[sheet_name]] if sheet_name else wb.worksheets

    titulos: List[Titulo] = []
    errors: List[str] = []

    for ws in sheets:
        found = _find_header_and_map(ws)
        if not found:
            continue

        header_row, colmap = found

        for excel_row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue

            payload = _row_to_titulo(row, colmap, ws.title)
            if payload is None:
                continue

            try:
                titulos.append(Titulo(**payload))
            except Exception as e:
                errors.append(f"{ws.title} linha {excel_row_idx}: {e}")

    if errors:
        raise DataFileError("Falha ao validar títulos:\n- " + "\n- ".join(errors))

    if not titulos:
        raise DataFileError(
            "Não encontrei nenhuma aba com colunas compatíveis (tipo/produto/título, vencimento e taxa).\n"
        )

    return titulos
